from __future__ import annotations

import csv
import io
import mimetypes
from pathlib import Path
from uuid import uuid4

from fastapi import HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.models.attachment import Attachment
from app.models.chat_message import ChatMessage
from app.models.thread import Thread
from app.models.user import User


class AttachmentService:
    _SUPPORTED_MIME_TYPES: dict[str, set[str]] = {
        "image": {"image/png", "image/jpeg", "image/jpg", "image/webp", "image/gif"},
        "video": {"video/mp4", "video/webm", "video/quicktime"},
        "table": {
            "text/csv",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
        "code": {
            "text/plain",
            "application/json",
            "text/markdown",
            "application/javascript",
            "text/javascript",
            "text/x-python",
        },
        "document": {
            "application/pdf",
            "application/msword",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/x-tex",
            "text/x-tex",
            "text/plain",
            "text/markdown",
        },
    }

    _EXTENSION_FILE_TYPES: dict[str, str] = {
        ".png": "image",
        ".jpg": "image",
        ".jpeg": "image",
        ".webp": "image",
        ".gif": "image",
        ".mp4": "video",
        ".webm": "video",
        ".mov": "video",
        ".csv": "table",
        ".xls": "table",
        ".xlsx": "table",
        ".py": "code",
        ".js": "code",
        ".ts": "code",
        ".tsx": "code",
        ".jsx": "code",
        ".java": "code",
        ".c": "code",
        ".cpp": "code",
        ".cs": "code",
        ".go": "code",
        ".rs": "code",
        ".sql": "code",
        ".json": "code",
        ".md": "document",
        ".txt": "document",
        ".tex": "document",
        ".pdf": "document",
        ".doc": "document",
        ".docx": "document",
    }

    @staticmethod
    async def _get_user_thread(db: AsyncSession, user: User, thread_id: int) -> Thread:
        thread = await db.scalar(select(Thread).where(Thread.id == thread_id, Thread.user_id == user.id))
        if not thread:
            raise HTTPException(status_code=404, detail="Thread not found")
        return thread

    @classmethod
    def _infer_file_type(cls, filename: str, mime_type: str) -> str:
        ext = Path(filename).suffix.lower()
        by_ext = cls._EXTENSION_FILE_TYPES.get(ext)
        if by_ext:
            return by_ext

        for file_type, mime_values in cls._SUPPORTED_MIME_TYPES.items():
            if mime_type in mime_values:
                return file_type

        raise HTTPException(status_code=400, detail="Unsupported file type")

    @classmethod
    def _validate_mime_type(cls, file_type: str, mime_type: str, filename: str) -> None:
        ext = Path(filename).suffix.lower()
        guessed_mime = mimetypes.guess_type(filename)[0]
        allowed = cls._SUPPORTED_MIME_TYPES.get(file_type, set())

        # Accept if declared MIME matches known allowed values for this file type.
        if mime_type in allowed:
            return

        # Fallback for browsers that send generic text/plain for source files.
        if file_type == "code" and mime_type == "text/plain" and ext in cls._EXTENSION_FILE_TYPES:
            return

        # Secondary fallback: guessed MIME matches the allowed list.
        if guessed_mime and guessed_mime in allowed:
            return

        raise HTTPException(status_code=400, detail="Unsupported or invalid MIME type")

    @staticmethod
    def _max_upload_bytes() -> int:
        return max(settings.MAX_UPLOAD_MB, 1) * 1024 * 1024

    async def upload(self, db: AsyncSession, user: User, thread_id: int, upload_file: UploadFile) -> Attachment:
        await self._get_user_thread(db, user, thread_id)

        if not upload_file.filename:
            raise HTTPException(status_code=400, detail="Filename is required")

        raw_bytes = await upload_file.read()
        if not raw_bytes:
            raise HTTPException(status_code=400, detail="Uploaded file is empty")

        if len(raw_bytes) > self._max_upload_bytes():
            raise HTTPException(
                status_code=413,
                detail=f"File exceeds MAX_UPLOAD_MB ({settings.MAX_UPLOAD_MB} MB)",
            )

        mime_type = (upload_file.content_type or "application/octet-stream").lower()
        file_type = self._infer_file_type(upload_file.filename, mime_type)
        self._validate_mime_type(file_type, mime_type, upload_file.filename)

        upload_root = Path(settings.UPLOAD_DIR).resolve()
        file_dir = upload_root / str(user.id) / str(thread_id)
        file_dir.mkdir(parents=True, exist_ok=True)

        safe_ext = Path(upload_file.filename).suffix.lower()
        stored_name = f"{uuid4().hex}{safe_ext}"
        file_path = file_dir / stored_name
        file_path.write_bytes(raw_bytes)

        item = Attachment(
            user_id=user.id,
            thread_id=thread_id,
            original_filename=upload_file.filename,
            file_path=str(file_path),
            mime_type=mime_type,
            file_type=file_type,
        )
        db.add(item)
        await db.commit()
        await db.refresh(item)
        return item

    async def resolve_for_chat(
        self,
        db: AsyncSession,
        user: User,
        thread_id: int,
        attachment_ids: list[int] | None,
    ) -> list[Attachment]:
        if not attachment_ids:
            return []

        result = await db.scalars(
            select(Attachment).where(
                Attachment.id.in_(attachment_ids),
                Attachment.user_id == user.id,
                Attachment.thread_id == thread_id,
            )
        )
        attachments = list(result)

        if len(attachments) != len(set(attachment_ids)):
            raise HTTPException(status_code=404, detail="One or more attachments were not found")

        return attachments

    @staticmethod
    def link_to_chat_message(attachments: list[Attachment], message: ChatMessage) -> None:
        for attachment in attachments:
            attachment.chat_message_id = message.id

    async def get_user_attachment(self, db: AsyncSession, user: User, attachment_id: int) -> Attachment:
        attachment = await db.scalar(
            select(Attachment).where(Attachment.id == attachment_id, Attachment.user_id == user.id)
        )
        if not attachment:
            raise HTTPException(status_code=404, detail="Attachment not found")
        return attachment

    @staticmethod
    def read_text_context(attachment: Attachment, max_chars: int = 4000) -> str:
        path = Path(attachment.file_path)
        if not path.exists():
            return f"[Missing attachment file: {attachment.original_filename}]"

        if attachment.file_type == "image":
            return ""

        if attachment.file_type == "video":
            # Video content is analysed visually via extracted frames in chat_service.
            # No text fallback is needed here.
            return ""

        raw_bytes = path.read_bytes()
        ext = path.suffix.lower()

        try:
            if ext == ".csv":
                return AttachmentService._extract_csv(raw_bytes, max_chars)
            if ext in {".xlsx", ".xls"}:
                return AttachmentService._extract_excel(raw_bytes, max_chars)
            if ext == ".pdf":
                return AttachmentService._extract_pdf(raw_bytes, max_chars)
            if ext == ".docx":
                return AttachmentService._extract_docx(raw_bytes, max_chars)

            decoded = raw_bytes.decode("utf-8", errors="ignore")
            return decoded[:max_chars]
        except Exception:
            return f"[Could not parse attachment text: {attachment.original_filename}]"

    @staticmethod
    def _extract_csv(raw_bytes: bytes, max_chars: int) -> str:
        text = raw_bytes.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        rows: list[str] = []
        for idx, row in enumerate(reader):
            rows.append(", ".join(col.strip() for col in row))
            if idx >= 30:
                break
        return "\n".join(rows)[:max_chars]

    @staticmethod
    def _extract_excel(raw_bytes: bytes, max_chars: int) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        output: list[str] = []
        for sheet in wb.worksheets[:2]:
            output.append(f"Sheet: {sheet.title}")
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                output.append(", ".join("" if value is None else str(value) for value in row))
                if idx >= 20:
                    break
        return "\n".join(output)[:max_chars]

    @staticmethod
    def _extract_pdf(raw_bytes: bytes, max_chars: int) -> str:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(raw_bytes))
        parts: list[str] = []
        for page in reader.pages[:5]:
            parts.append(page.extract_text() or "")
        return "\n".join(parts)[:max_chars]

    @staticmethod
    def _extract_docx(raw_bytes: bytes, max_chars: int) -> str:
        from docx import Document

        doc = Document(io.BytesIO(raw_bytes))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return text[:max_chars]


attachment_service = AttachmentService()
